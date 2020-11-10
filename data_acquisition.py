#!/usr/bin/env python3
#####################################
#    LAST UPDATED     14 OCT 2020   #
#####################################
"""
Historical script of how I got the spreadsheet dataset of NFL Overtime games

The website converted to a subscription model, and the information is now behind a paywall. This script worked
at the time, but it may no longer work without a subscription to stathead.com
"""
import openpyxl
import bs4
import requests


def get_all_overtimes():
    """
    Get all overtime games and return them as a list
    :return: list of all overtime games
    """
    headers = {'user-agent':
                       "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:74.0) Gecko/20100101 Firefox/74.0"}
    
    master_list = []
    for number in [0, 100, 200]:
        
        url = 'https://www.pro-football-reference.com/play-index/tgl_finder.cgi?request=1&match=game&year_min=' \
              '2012&year_max=2020&game_type=E&game_num_min=0&game_num_max=99&week_num_min=0&week_num_max=99' \
              '&temperature_gtlt=lt&overtime=Y&c5val=1.0&order_by=game_date&offset={}'.format(number)
        
        text = requests.get(url, headers=headers).text
        
        soup = bs4.BeautifulSoup(text, 'lxml')
        
        table = soup.find_all('table', class_='sortable stats_table')
        
        for tablerow in table:
            for another_row in tablerow:
                if isinstance(another_row, bs4.element.Tag):
                    for row in another_row:
                        if isinstance(row, bs4.element.Tag):
                            mini_list = []
                            for cell in row:
                                if isinstance(cell, bs4.element.Tag):
                                    mini_list.append(cell.text)
                                if mini_list not in master_list and len(mini_list) == 13:
                                    master_list.append(mini_list)
                                    
    return master_list


def get_coin_toss_winner(overtimes):
    """
    Try to build the URL for the individual game and scrape that page for the Overtime Coin Toss winner
    :param overtimes: list of all overtime games
    :return: list with new column "overtime coin toss winner" and "won coin toss" added
    """
    headers = {'user-agent':
                           "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:74.0) Gecko/20100101 Firefox/74.0"}
        
    for index, line in enumerate(overtimes):
        if index == 0:
            continue
        else:
            rank = line[0]
            team = line[1]
            year = line[2]
            date = line[3]
            time = line[4]
            localtime = line[5]
            at = line[6]
            opponent = line[7]
            weeknum = line[8]
            gamenum = line[9]
            day = line[10]
            resultscore = line[11]
            ot = line[12]
            value = ''
            ot_coin_toss_winner = 'UNK'
            
            if line[6]:
                another_conv_dict = {
                    'LAC': 'SDG',
                    'HOU': 'HTX',
                    'ARI': 'CRI',
                    'OAK': 'ORI',
                    'TEN': 'OTI'}
                if opponent in another_conv_dict.keys():
                    new_opponent = another_conv_dict[opponent]
                else:
                    new_opponent = opponent
                    
                url = 'https://www.pro-football-reference.com/boxscores/{}0{}.htm'.format(date.replace('-', ''),
                                                                                          new_opponent.lower())
                text = requests.get(url, headers=headers).text
                
                if 'Page Not Found (404 error)' not in text:
                    first_index = text.index('Won OT Toss')+54
                    second_index = text[first_index:].index('<')
                    ot_coin_toss_winner = text[first_index:first_index+second_index].strip()
                    conv_dict = {
                        'Vikings': 'MIN',
                        'Jaguars': 'JAX',
                        'Titans': 'TEN',
                        'Jets': 'NYJ',
                        'Dolphins': 'MIA',
                        'Saints': 'NOR',
                        'Chiefs': 'KAN',
                        'Lions': 'DET',
                        'Cardinals': 'ARI',
                        'Bills': 'BUF',
                        'Raiders': 'OAK/LAS',
                        'Patriots': 'NWE',
                        'Colts': 'IND',
                        'Rams': 'STL/LAR',
                        '49ers': 'SFO',
                        'Steelers': 'PIT',
                        'Buccaneers': 'TAM',
                        'Panthers': 'CAR',
                        'Cowboys': 'DAL',
                        'Browns': 'CLE',
                        'Texans': 'HOU',
                        'Chargers': 'SDG/LAC',
                        'Ravens': 'BAL',
                        'Seahawks': 'SEA',
                        'Bears': 'CHI',
                        'Redskins': 'WAS',
                        'Bengals': 'CIN',
                        'Packers': 'GNB',
                        'Broncos': 'DEN',
                        'Falcons': 'ATL',
                        'Eagles': 'PHI',
                        'Giants': 'NYG',
                        }
                    
                    if ot_coin_toss_winner in conv_dict.keys():
                        if '/' in conv_dict[ot_coin_toss_winner]:
                            winner = conv_dict[ot_coin_toss_winner]
                            city1, city2 = conv_dict[ot_coin_toss_winner].split('/')
                            if winner == city1 or winner == city2:
                                value = True
                            else:
                                #print('{} {}, {}, {} {} {}'.format(rank, ot_coin_toss_winner, winner,
                                # city1, city2, team))
                                value = False
                        else:
                            if conv_dict[ot_coin_toss_winner] == team:
                                value = True
                            else:
                                #print('{}, {}, {}, {}'.format(rank, ot_coin_toss_winner,
                                # conv_dict[ot_coin_toss_winner], team))
                                value = False
                    else:
                        print('{} {} not in dict'.format(rank, ot_coin_toss_winner))
                        
                else:
                    print('{} Wrong url for {} and {}'.format(rank, date, opponent))
            
            try:
                overtimes[index] = [rank, team, year, date, time, localtime, at, opponent,
                                    weeknum, gamenum, day, resultscore, ot, ot_coin_toss_winner, value]
            except NameError:
                overtimes[index] = [rank, team, year, date, time, localtime, at, opponent,
                                    weeknum, gamenum, day, resultscore, ot, 'UNK', value]
                
    return overtimes


def save_to_excel(overtimes):
    """
    Save the list to a workbook
    :param overtimes: list of overtime games
    :return: None
    """
    workbook = openpyxl.Workbook()
    ws1 = workbook.active
    
    row_counter = 1
    col_counter = 1
    for list_row in overtimes:
        for list_cell in list_row:
            ws1.cell(row=row_counter, column=col_counter).value = list_cell
            col_counter += 1
        row_counter += 1
        col_counter = 1
        
    workbook.save('/Users/Alex/Desktop/overtimes.xlsx')
    
    
def verify_xlsx(path):
    """
    Function helped troubleshoot to ensure that each game had an overtime winner and overtime loser
    x = verify_xlsx('/Users/Alex/Documents/Python3/nfl_overtime/overtimes_clean.xlsx')

    for key, val in x.items():
        val_true = val[True]
        val_false = val[False]

        if val_true != val_false:
            print(key)

    :param path: Str path to spreadsheet
    :return: None
    """
    row_counter = 2
    winnercol = 15
    datecol = 4
    return_dict = {}
    
    workbook = openpyxl.load_workbook(path)
    ws = workbook["Sheet"]
    
    while ws.cell(row=row_counter, column=datecol).value:
        date = ws.cell(row=row_counter, column=datecol).value
        ot_winner = ws.cell(row=row_counter, column=winnercol).value
        
        if date in return_dict.keys():
            if ot_winner:
                return_dict[date][True] += 1
            else:
                return_dict[date][False] += 1
        else:
            if ot_winner:
                return_dict[date] = {True: 1, False: 0}
            else:
                return_dict[date] = {True: 0, False: 1}
                
        row_counter += 1

    return return_dict


overtime = get_all_overtimes()
    
overtime[0] = ['Rank',
                 'Team',
                 'Year',
                 'Date',
                 'Time',
                 'LocalTime',
                 'at',
                 'Opponent',
                 'Week Number',
                 'Game Number',
                 'Day',
                 'Result Score',
                 'OT', 'OT Coin toss winner', 'Won OT toss']

overtime = get_coin_toss_winner(overtime)
